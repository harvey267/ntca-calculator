import argparse
import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# Default segment rates back-solved from verified NTCA Pension Point outputs (Apr 2026).
# These are the §417(e) HQM corporate bond segment rates NTCA applies for the 2026 plan year.
# Update annually when NTCA refreshes rates (typically August look-back month).
DEFAULT_SEG_RATES = {"seg1": 0.0410, "seg2": 0.0520, "seg3": 0.0580}

MORTALITY_TABLE_DIR = os.path.join(os.path.dirname(__file__), "ntca_pension_dashboard")
DEFAULT_MORTALITY_CSV = os.path.join(MORTALITY_TABLE_DIR, "n-25-40.csv")  # IRS Notice 2025-40 (2026 plan year)


def parse_args():
    parser = argparse.ArgumentParser(
        description="NTCA §417(e) lump sum calculator — computes present value of pension annuity",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 ntca_lump_sum_calculator.py \\
    --birth-date 1970-01-01 \\
    --retirement-date 2030-06-01 \\
    --salary 80000 \\
    --years 25 \\
    --benefit-rate 0.019

Finding your inputs:
  Benefit rate    — Your employer's elected rate from the NTCA Adoption Agreement.
                    Common values: 0.019 (1.9%), 0.017 (1.7%), 0.015 (1.5%).
                    Check your Summary Plan Description or ask your HR/benefits contact.
  High-5 salary   — Average of your 5 highest compensation years in the last 10 plan years.
  Years of service — Total years of NTCA plan participation (fractions count, e.g. 25.833).
  Segment rates   — IRS §417(e) HQM corporate bond rates. NTCA uses August rates with a
                    one-year look-back. Defaults are verified against 2026 NTCA calculations.
        """
    )
    parser.add_argument("--birth-date",      required=True,
                        help="Date of birth (YYYY-MM-DD)")
    parser.add_argument("--retirement-date", required=True,
                        help="Planned benefit commencement date (YYYY-MM-DD)")
    parser.add_argument("--salary",          required=True, type=float,
                        help="High-5 average salary ($) — average of 5 highest years in last 10")
    parser.add_argument("--years",           required=True, type=float,
                        help="Years of NTCA plan participation (fractions valid, e.g. 25.833)")
    parser.add_argument("--benefit-rate",    required=True, type=float,
                        help="Your employer's elected benefit rate (e.g. 0.019 for 1.9%%). "
                             "Find this in your Summary Plan Description or Adoption Agreement. "
                             "Not the same for all NTCA member employers.")
    parser.add_argument("--seg1",            type=float, default=None,
                        help=f"Segment 1 rate override (0-5yr, decimal). Default: {DEFAULT_SEG_RATES['seg1']}")
    parser.add_argument("--seg2",            type=float, default=None,
                        help=f"Segment 2 rate override (5-20yr, decimal). Default: {DEFAULT_SEG_RATES['seg2']}")
    parser.add_argument("--seg3",            type=float, default=None,
                        help=f"Segment 3 rate override (20yr+, decimal). Default: {DEFAULT_SEG_RATES['seg3']}")
    parser.add_argument("--mortality-csv",   default=DEFAULT_MORTALITY_CSV,
                        help="Path to IRS §417(e) mortality table CSV (Age, 417(e) columns). "
                             f"Default: {DEFAULT_MORTALITY_CSV}")
    parser.add_argument("--output",          default="ntca_lump_sum.xlsx",
                        help="Output Excel filename (default: ntca_lump_sum.xlsx)")
    return parser.parse_args()


def load_mortality_table(csv_path):
    """
    Loads IRS §417(e) mortality table from CSV (columns: Age, 417(e)).
    The 417(e) column is the annual mortality rate (qx).
    Returns (monthly_ages array, monthly_conditional_survival function).
    """
    df = pd.read_csv(csv_path)
    df.columns = ["age", "qx"]
    df = df.dropna()
    ages = df["age"].values
    qx   = df["qx"].values

    # Build cumulative survival: S(n) = prod(1 - q(i)) for i < n
    survival = np.ones(len(ages))
    for i in range(1, len(ages)):
        survival[i] = survival[i - 1] * (1 - qx[i - 1])

    monthly_ages     = np.arange(0, ages[-1] + 1 / 12, 1 / 12)
    monthly_survival = np.interp(monthly_ages, ages, survival)
    return monthly_ages, monthly_survival


def ntca_birth_date(birth_date):
    # NTCA snaps age to the 1st of the birth month for all calculations.
    # A participant born June 10 is treated as born June 1.
    return birth_date.replace(day=1)


def calculate_exact_age(birth_date, retirement_date):
    snapped = ntca_birth_date(birth_date)
    delta = retirement_date - snapped
    return round(delta.days / 365.25, 4)


def calculate_monthly_annuity(salary, years, benefit_rate):
    return round((benefit_rate * salary * years) / 12, 2)


def build_annuity_stream(start_age, monthly_annuity, segment_rates, monthly_ages, monthly_survival):
    """
    Builds 65-year monthly annuity stream using conditional survival probabilities
    and §417(e) three-segment discount rates.
    """
    s_at_start = np.interp(start_age, monthly_ages, monthly_survival)
    rows = []
    for month in range(1, 781):
        age = start_age + month / 12
        segment = (
            segment_rates["seg1"] if month <= 60 else
            segment_rates["seg2"] if month <= 240 else
            segment_rates["seg3"]
        )
        discount_factor  = 1 / ((1 + segment / 12) ** month)
        cond_survival    = np.interp(age, monthly_ages, monthly_survival) / s_at_start
        pv               = monthly_annuity * cond_survival * discount_factor
        rows.append([month, round(age, 4), monthly_annuity, round(cond_survival, 6),
                     segment, round(discount_factor, 6), round(pv, 2)])

    return pd.DataFrame(rows, columns=[
        "Month", "Age", "Annuity", "Survival (conditional)",
        "Segment Rate", "Discount Factor", "PV of Payment"
    ])


def calculate_lump_sum(df_stream):
    full_lump  = df_stream["PV of Payment"].sum()
    pv_factor  = full_lump / (df_stream["Annuity"].iloc[0] * 12)
    return round(full_lump, 2), round(pv_factor, 4)


def export_to_excel(inputs, df_stream, summary, filename):
    wb = Workbook()

    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    for key, value in inputs.items():
        ws_inputs.append([key, value])

    ws_stream = wb.create_sheet("Annuity Stream")
    for r in dataframe_to_rows(df_stream, index=False, header=True):
        ws_stream.append(r)

    ws_summary = wb.create_sheet("Summary")
    for key, value in summary.items():
        ws_summary.append([key, value])

    wb.save(filename)
    print(f"Exported to {filename}")


if __name__ == "__main__":
    args = parse_args()

    birth_date      = datetime.strptime(args.birth_date, "%Y-%m-%d")
    retirement_date = datetime.strptime(args.retirement_date, "%Y-%m-%d")

    segment_rates = {
        "seg1": args.seg1 if args.seg1 is not None else DEFAULT_SEG_RATES["seg1"],
        "seg2": args.seg2 if args.seg2 is not None else DEFAULT_SEG_RATES["seg2"],
        "seg3": args.seg3 if args.seg3 is not None else DEFAULT_SEG_RATES["seg3"],
    }

    print(f"Loading mortality table from {os.path.basename(args.mortality_csv)}...")
    monthly_ages, monthly_survival = load_mortality_table(args.mortality_csv)

    print("Calculating...")
    exact_age       = calculate_exact_age(birth_date, retirement_date)
    monthly_annuity = calculate_monthly_annuity(args.salary, args.years, args.benefit_rate)
    df_stream       = build_annuity_stream(exact_age, monthly_annuity, segment_rates,
                                           monthly_ages, monthly_survival)
    lump_sum, pv_factor = calculate_lump_sum(df_stream)

    inputs = {
        "Birth Date":         birth_date.strftime("%Y-%m-%d"),
        "Retirement Date":    retirement_date.strftime("%Y-%m-%d"),
        "Exact Age":          exact_age,
        "High-5 Salary":      f"${args.salary:,.2f}",
        "Years of Service":   args.years,
        "Benefit Rate":       f"{args.benefit_rate:.1%}",
        "Monthly Annuity":    f"${monthly_annuity:,.2f}",
        "Segment 1 (0-5yr)":  f"{segment_rates['seg1']:.2%}",
        "Segment 2 (5-20yr)": f"{segment_rates['seg2']:.2%}",
        "Segment 3 (20yr+)":  f"{segment_rates['seg3']:.2%}",
        "Mortality Table":    os.path.basename(args.mortality_csv),
    }
    summary = {
        "Full Lump Sum":  f"${lump_sum:,.2f}",
        "PV Factor":      pv_factor,
        "Monthly Annuity": f"${monthly_annuity:,.2f}",
        "Annual Annuity": f"${monthly_annuity * 12:,.2f}",
    }

    print(f"\nBenefit rate:     {args.benefit_rate:.1%}")
    print(f"Monthly annuity:  ${monthly_annuity:,.2f}")
    print(f"Annual annuity:   ${monthly_annuity * 12:,.2f}")
    print(f"Full lump sum:    ${lump_sum:,.2f}")
    print(f"PV factor:        {pv_factor}")
    print(f"Segment rates:    {segment_rates['seg1']:.2%} / {segment_rates['seg2']:.2%} / {segment_rates['seg3']:.2%}")

    export_to_excel(inputs, df_stream, summary, args.output)
